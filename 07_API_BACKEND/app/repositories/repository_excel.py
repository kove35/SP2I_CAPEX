from __future__ import annotations

import csv
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class RepositoryExcel:
    """
    Lecture Excel isolee du reste du backend.

    Le service IA recoit des donnees tabulaires simples et ne connait pas
    `openpyxl`. Cela permettra de remplacer le lecteur plus tard si besoin.
    """

    def lire_workbook(self, contenu: bytes, nom_fichier: str = "workbook.xlsx") -> dict[str, list[list[Any]]]:
        extension = Path(nom_fichier).suffix.lower()

        if extension == ".csv":
            return {"CSV": self._lire_csv(contenu)}

        if extension == ".xls":
            return self._lire_xls(contenu)

        workbook = load_workbook(BytesIO(contenu), read_only=True, data_only=True)
        feuilles: dict[str, list[list[Any]]] = {}

        for worksheet in workbook.worksheets:
            lignes: list[list[Any]] = []
            for row in worksheet.iter_rows(values_only=True):
                lignes.append(list(row))
            feuilles[worksheet.title] = lignes

        return feuilles

    def _lire_csv(self, contenu: bytes) -> list[list[Any]]:
        """
        Lit un CSV sans hypothese forte sur le separateur.

        Les exports metier alternent souvent entre `;` et `,`. On laisse Python
        detecter le dialecte puis on garde une structure identique a une feuille
        Excel pour ne pas changer les services existants.
        """
        texte = contenu.decode("utf-8-sig", errors="replace")
        echantillon = texte[:2048]
        try:
            dialect = csv.Sniffer().sniff(echantillon, delimiters=";,	")
            return [list(row) for row in csv.reader(texte.splitlines(), dialect)]
        except csv.Error:
            return [list(row) for row in csv.reader(texte.splitlines(), delimiter=";")]

    def _lire_xls(self, contenu: bytes) -> dict[str, list[list[Any]]]:
        """
        Lit les anciens fichiers Excel `.xls` via pandas/xlrd.

        L'import est volontairement local a cette methode : si un deploiement
        n'a pas besoin du format historique `.xls`, le reste du backend continue
        de fonctionner normalement.
        """
        try:
            import pandas as pd
        except ImportError as erreur:
            raise RuntimeError("Le support .xls necessite pandas.") from erreur

        try:
            workbook = pd.read_excel(BytesIO(contenu), sheet_name=None, header=None)
        except ImportError as erreur:
            raise RuntimeError("Le support .xls necessite la dependance xlrd.") from erreur

        feuilles: dict[str, list[list[Any]]] = {}
        for sheet_name, dataframe in workbook.items():
            dataframe = dataframe.where(pd.notnull(dataframe), None)
            feuilles[str(sheet_name)] = dataframe.values.tolist()

        return feuilles
