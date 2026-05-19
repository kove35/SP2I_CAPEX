from __future__ import annotations

import hashlib
import json
from io import BytesIO
from pathlib import Path
from time import perf_counter
from typing import Any

from sqlalchemy.orm import Session

from app.analytics.cache import analytics_cache
from app.analytics.repositories import AnalyticsRepository
from app.analytics.schemas import AnalyticsQuery
from app.analytics.utils.display_text import normalize_payload_labels


RACINE = Path(__file__).resolve().parents[4]

SUPPORTED_CURRENCIES = {
    "FCFA": {"label": "Franc CFA", "symbol": "FCFA", "to_fcfa": 1.0},
    "USD": {"label": "Dollar americain", "symbol": "$", "to_fcfa": 610.0},
    "EUR": {"label": "Euro", "symbol": "EUR", "to_fcfa": 655.957},
}

CHINA_PORTS = ["Ningbo", "Shanghai", "Shenzhen", "Guangzhou"]

CHINA_SUPPLIERS = [
    {
        "supplier": "Shanghai Aludream Building Material Co., Ltd.",
        "category": "Alucobond",
        "port": "Shanghai",
        "fob_usd": 18.2,
        "moq": "1 500 m2",
        "lead_time_days": 38,
        "quality": 88,
        "certifications": ["ISO 9001", "CE"],
        "history": "Fabricant ACP reference public web",
        "reliability": 86,
        "capacity": "45 000 m2/mois",
        "risk": 34,
        "source_url": "https://www.aludreambond.com/",
        "source_confidence": 0.82,
    },
    {
        "supplier": "Shanghai FATO Group Co., Ltd.",
        "category": "Electricite",
        "port": "Shanghai",
        "fob_usd": 42.5,
        "moq": "500 unites",
        "lead_time_days": 45,
        "quality": 91,
        "certifications": ["IEC", "CE", "UL", "ISO 14001"],
        "history": "Fabricant appareillage basse tension",
        "reliability": 89,
        "capacity": "12 containers/mois",
        "risk": 31,
        "source_url": "https://www.sh-fato.com/",
        "source_confidence": 0.86,
    },
    {
        "supplier": "Midea Group",
        "category": "CVC",
        "port": "Guangzhou",
        "fob_usd": 310.0,
        "moq": "80 equipements",
        "lead_time_days": 52,
        "quality": 87,
        "certifications": ["CE", "UL"],
        "history": "Groupe industriel HVAC public",
        "reliability": 88,
        "capacity": "600 equipements/mois",
        "risk": 36,
        "source_url": "https://www.midea.com/",
        "source_confidence": 0.78,
    },
    {
        "supplier": "Shenzhen Sanitary Technology Benchmark",
        "category": "Sanitaires",
        "port": "Shenzhen",
        "fob_usd": 56.0,
        "moq": "300 kits",
        "lead_time_days": 34,
        "quality": 84,
        "certifications": ["WRAS", "ISO 9001"],
        "history": "Reference categorie a confirmer par RFQ",
        "reliability": 72,
        "capacity": "8 000 kits/mois",
        "risk": 48,
        "source_url": "",
        "source_confidence": 0.55,
    },
    {
        "supplier": "ZMR Windows and Doors",
        "category": "Menuiserie aluminium",
        "port": "Guangzhou",
        "fob_usd": 96.0,
        "moq": "1 container",
        "lead_time_days": 48,
        "quality": 90,
        "certifications": ["CE", "National inspection"],
        "history": "Fabricant portes/fenetres aluminium Foshan",
        "reliability": 84,
        "capacity": "20 000 ml/mois",
        "risk": 37,
        "source_url": "https://zmrdoor.com/",
        "source_confidence": 0.76,
    },
    {
        "supplier": "SJEC Corporation",
        "category": "Ascenseurs",
        "port": "Shanghai",
        "fob_usd": 28500.0,
        "moq": "1 appareil",
        "lead_time_days": 78,
        "quality": 92,
        "certifications": ["EN81", "CE", "ISO 9001", "ISO 14001"],
        "history": "Fabricant ascenseurs Suzhou reference CCCME",
        "reliability": 88,
        "capacity": "40 appareils/mois",
        "risk": 38,
        "source_url": "https://www.cccme.cn/shop/cccme14022/index.aspx",
        "source_confidence": 0.88,
    },
    {
        "supplier": "Guangzhou Medivara Medical Co., Ltd.",
        "category": "Mobilier medical",
        "port": "Guangzhou",
        "fob_usd": 145.0,
        "moq": "250 unites",
        "lead_time_days": 41,
        "quality": 86,
        "certifications": ["ISO 13485", "CE"],
        "history": "Mobilier medical et lits hospitaliers",
        "reliability": 80,
        "capacity": "5 000 unites/mois",
        "risk": 36,
        "source_url": "https://www.medicaldevicez.com/",
        "source_confidence": 0.74,
    },
]

CHINA_SCENARIOS = [
    {"code": "eco_china", "label": "Eco China", "discount": 0.21, "quality": 78, "risk": 62, "lead_time": 70, "security": 0.035},
    {"code": "standard_china", "label": "Standard China", "discount": 0.165, "quality": 86, "risk": 42, "lead_time": 58, "security": 0.025},
    {"code": "premium_china", "label": "Premium China", "discount": 0.105, "quality": 93, "risk": 28, "lead_time": 48, "security": 0.018},
]

LOCAL_SUPPLIERS = {
    "electricite": {"supplier": "Societe Electrique Congo", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 12, "quality": 82, "history": "Fourniture locale recurrente"},
    "menuiserie": {"supplier": "Ateliers Aluminium Pointe-Noire", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 18, "quality": 80, "history": "Fabrication et pose locale"},
    "cvc": {"supplier": "Congo Clim Services", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 21, "quality": 78, "history": "Maintenance CVC locale"},
    "climatisation": {"supplier": "Congo Clim Services", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 21, "quality": 78, "history": "Maintenance CVC locale"},
    "plomberie": {"supplier": "Sanibat Congo", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 14, "quality": 79, "history": "Reseaux sanitaires et plomberie"},
    "ascenseur": {"supplier": "Congo Lift Maintenance", "city": "Brazzaville", "country": "Congo", "lead_time_days": 45, "quality": 76, "history": "Maintenance et coordination installation"},
    "alucobond": {"supplier": "Facade Congo SARL", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 24, "quality": 77, "history": "Pose facade et habillage"},
    "default": {"supplier": "Fournisseur local a qualifier", "city": "Pointe-Noire", "country": "Congo", "lead_time_days": 15, "quality": 74, "history": "A confirmer par consultation locale"},
}

IMPORT_RISKS = [
    {"label": "Maritime", "probability": 0.44, "impact_rate": 0.018, "action": "Reserver capacite et consolider containers."},
    {"label": "Douane", "probability": 0.35, "impact_rate": 0.013, "action": "Preparer nomenclatures HS et documents origine."},
    {"label": "Devise", "probability": 0.58, "impact_rate": 0.024, "action": "Figer USD/EUR ou integrer une clause de change."},
    {"label": "Fournisseur", "probability": 0.28, "impact_rate": 0.016, "action": "Audit fournisseur et inspection pre-embarquement."},
    {"label": "Qualite", "probability": 0.24, "impact_rate": 0.014, "action": "Echantillons, fiches techniques et controle tiers."},
    {"label": "Geopolitique", "probability": 0.22, "impact_rate": 0.011, "action": "Plan B fournisseur et marge planning."},
    {"label": "Congestion portuaire", "probability": 0.31, "impact_rate": 0.012, "action": "Choisir port alternatif Ningbo/Shenzhen selon lot."},
]


class AnalyticsService:
    """Service applicatif du moteur BI proprietaire SP2I."""

    def __init__(self, db: Session) -> None:
        self.repository = AnalyticsRepository(db)

    def dashboard(self, query: AnalyticsQuery, dashboard_type: str = "direction") -> dict[str, Any]:
        return self._cached(f"dashboard:{dashboard_type}", query, lambda: self._build_dashboard(query, dashboard_type))

    def capex(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("capex", query, lambda: self._build_dashboard(query, "capex"))

    def kpis(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("kpis", query, lambda: self._response(query, kpis=self.repository.kpis(query)))

    def risk(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("risk", query, lambda: self._response(query, charts={"risk_matrix": self.repository.risk_matrix(query)}))

    def procurement(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("procurement", query, lambda: self._build_dashboard(query, "procurement"))

    def gain_analysis(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("gain-analysis", query, lambda: self._build_gain_analysis(query))

    def suppliers(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("suppliers", query, lambda: self._build_suppliers(query))

    def procurement_lines(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("procurement-lines", query, lambda: self._build_procurement_lines(query))

    def procurement_scenarios(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("procurement-scenarios", query, lambda: self._build_procurement_scenarios(query))

    def currency(self, query: AnalyticsQuery) -> dict[str, Any]:
        active = self._active_currency(query)
        return self._response(
            query,
            kpis={"active_currency": active, "usd_to_fcfa": 610, "eur_to_fcfa": 655.957},
            charts={
                "currencies": [
                    {"code": code, **metadata}
                    for code, metadata in SUPPORTED_CURRENCIES.items()
                ],
                "fx_risk": [
                    {"variation": "USD +5%", "gain_impact_rate": -0.045},
                    {"variation": "USD +10%", "gain_impact_rate": -0.09},
                    {"variation": "EUR +5%", "gain_impact_rate": -0.032},
                ],
            },
            metadata={"engine": "SP2I Currency Engine V1", "source": "SP2I_REFERENCE"},
        )

    def import_risks(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("import-risks", query, lambda: self._build_import_risks(query))

    def logistics(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._cached("logistics", query, lambda: self._build_logistics_plan(query))

    def scenarios(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._response(query, table=self.repository.scenarios())

    def heatmap(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._response(query, charts={"heatmap": self.repository.heatmap(query)})

    def drilldown(self, query: AnalyticsQuery) -> dict[str, Any]:
        path = self.repository.drilldown_path(query.drilldown_level)
        grouped = self.repository.grouped(query, default_group=path["next"] or path["current"])
        return self._response(query, charts={"drilldown": grouped}, metadata={"drilldown": path})

    def timeline(self, query: AnalyticsQuery) -> dict[str, Any]:
        return self._response(query, charts={"timeline": self.repository.timeline(query)})

    def filter_options(self) -> dict[str, Any]:
        options = self.repository.filter_options()
        return normalize_payload_labels({
            "batiments": options.get("batiments", []),
            "niveaux": options.get("niveaux", []),
            "lots": options.get("lots", []),
            "familles": options.get("familles", []),
            "import_local": options.get("import_local", []),
        })

    def system_health(self) -> dict[str, Any]:
        return {
            "status": "SUCCESS",
            "filters": {},
            "pagination": {},
            "kpis": {},
            "charts": {},
            "table": [],
            "metadata": {
                "engine": "SP2I Analytics Engine V1",
                "database": "configured",
                "cache": analytics_cache.status(),
            },
        }

    def cache_status(self) -> dict[str, Any]:
        return {
            "status": "SUCCESS",
            "filters": {},
            "pagination": {},
            "kpis": {},
            "charts": {},
            "table": [],
            "metadata": analytics_cache.status(),
        }

    def debug_pipeline(self) -> dict[str, Any]:
        debug = self.repository.pipeline_debug()
        latest_source = self._latest_pipeline_source()
        return {
            "status": "SUCCESS",
            "filters": {},
            "pagination": {"page": 1, "page_size": len(debug["preview"]), "total": debug["fact_metre_count"]},
            "kpis": {
                "capex_brut": round(float(debug["sums"].get("capex_local_total") or 0), 2),
                "capex_optimise": round(float(debug["sums"].get("capex_optimise_total") or 0), 2),
                "economie_nette": round(float(debug["sums"].get("economie_total") or 0), 2),
                "nb_lignes": debug["fact_metre_count"],
            },
            "charts": {},
            "table": debug["preview"],
            "metadata": {
                "engine": "SP2I Analytics Engine V1",
                "debug": {
                    "columns": debug["columns"],
                    "sums": debug["sums"],
                    "views": debug["views"],
                    "latest_source": latest_source,
                    "cache": analytics_cache.status(),
                },
            },
            "warnings": debug["warnings"],
        }

    def qa_summary(self) -> dict[str, Any]:
        """Synthese QA lisible pour valider rapidement le moteur analytics."""
        query = AnalyticsQuery()
        kpis = self.repository.kpis(query)
        table, total = self.repository.table(query)
        heatmap = self.repository.heatmap(query)
        timeline = self.repository.timeline(query)
        grouped = self.repository.grouped(query, default_group="lot")
        cache_status = analytics_cache.status()

        capex_brut = float(kpis.get("capex_brut") or 0)
        economie = float(kpis.get("economie_nette") or 0)
        nb_lignes = int(kpis.get("nb_lignes") or 0)
        lots = {str(row.get("label")) for row in grouped if row.get("label")}
        checks = {
            "postgresql_connected": True,
            "fact_metre_non_empty": nb_lignes > 0,
            "nb_lignes_gt_500": nb_lignes > 500,
            "capex_brut_gt_0": capex_brut > 0,
            "economie_gt_0": economie > 0,
            "lots_gt_0": len(lots) > 0,
            "charts_ready": bool(grouped and heatmap),
            "ag_grid_ready": bool(table),
            "timeline_ready": bool(timeline),
            "cache_ready": cache_status.get("backend") == "in-memory",
        }
        warnings = [
            f"Check KO: {name}"
            for name, ok in checks.items()
            if not ok
        ]

        return {
            "status": "SUCCESS",
            "filters": {},
            "pagination": {"page": 1, "page_size": len(table), "total": total},
            "kpis": kpis,
            "charts": {
                "bar": grouped,
                "heatmap": heatmap,
                "timeline": timeline,
            },
            "table": table,
            "metadata": {
                "engine": "SP2I Analytics Engine V1",
                "qa_status": "PASS" if not warnings else "WARN",
                "checks": checks,
                "cache": cache_status,
                "dataset_target": {
                    "name": "SP2I_CAPEX_DEMO_V1",
                    "expected_rows": 584,
                    "expected_capex": 1129667152,
                },
            },
            "warnings": warnings,
        }

    def data_quality(self) -> dict[str, Any]:
        """Controle qualite bout-en-bout Excel -> FACT_METRE -> cockpit."""
        query = AnalyticsQuery()
        debug = self.repository.pipeline_debug()
        metrics = self.repository.quality_metrics()
        source = self._latest_pipeline_source()
        history = self.repository.import_audit_history()
        source_rows = int(source.get("rows_in_source_json") or 0)
        source_capex = float(source.get("capex_source") or 0)
        fact_rows = int(metrics.get("nb_lignes") or debug.get("fact_metre_count") or 0)
        analytics_capex = float(metrics.get("capex_local_total") or 0)
        capex_delta = analytics_capex - source_capex
        capex_delta_pct = abs(capex_delta) / source_capex if source_capex else 0
        line_delta = fact_rows - source_rows
        line_delta_pct = abs(line_delta) / source_rows if source_rows else 0
        family_pending = int(metrics.get("lignes_famille_a_classer") or 0)
        invalid_rows = (
            int(metrics.get("lignes_quantite_invalide") or 0)
            + int(metrics.get("lignes_capex_invalide") or 0)
            + int(metrics.get("lignes_sans_lot") or 0)
            + int(metrics.get("lignes_sans_designation") or 0)
        )
        anomalies = list(source.get("ai_anomalies") or [])
        warnings: list[str] = []
        if not source.get("available"):
            warnings.append("Aucun fichier source pipeline disponible.")
        if capex_delta_pct > 0.005:
            warnings.append("Ecart financier superieur a la tolerance de 0,5 %.")
        if source_rows and fact_rows != source_rows:
            warnings.append("Le nombre de lignes source et FACT_METRE differe.")
        if family_pending:
            warnings.append(f"{family_pending} lignes restent sans classification metier robuste.")
        if invalid_rows:
            warnings.append(f"{invalid_rows} controles ligne sont en anomalie dans FACT_METRE.")

        score = 100.0
        score -= min(35, capex_delta_pct * 1000)
        score -= min(20, line_delta_pct * 100)
        score -= min(20, (family_pending / max(fact_rows, 1)) * 100)
        score -= min(15, (invalid_rows / max(fact_rows, 1)) * 100)
        score -= min(10, len(anomalies) * 1.5)
        score = round(max(0, score), 1)

        checks = {
            "excel_source_available": bool(source.get("available")),
            "financial_reconciliation_ok": capex_delta_pct <= 0.005,
            "pipeline_rows_ok": bool(source_rows and fact_rows == source_rows),
            "fact_metre_non_empty": fact_rows > 0,
            "taxonomy_ok": family_pending == 0,
            "line_quality_ok": invalid_rows == 0,
        }

        return normalize_payload_labels({
            "status": "SUCCESS",
            "filters": {},
            "pagination": {"page": 1, "page_size": len(anomalies), "total": len(anomalies)},
            "kpis": {
                "score_qualite": score,
                "capex_source": round(source_capex, 2),
                "capex_analytics": round(analytics_capex, 2),
                "ecart_capex": round(capex_delta, 2),
                "ecart_capex_pct": round(capex_delta_pct, 6),
                "lignes_excel": source_rows,
                "lignes_fact_metre": fact_rows,
                "lignes_rejetees": max(source_rows - fact_rows, 0),
                "lignes_famille_a_classer": family_pending,
                "anomalies": len(anomalies) + invalid_rows,
            },
            "charts": {
                "pipeline": [
                    {"label": "Excel source", "value": source_rows},
                    {"label": "Parsing IA", "value": source_rows},
                    {"label": "FACT_METRE", "value": fact_rows},
                    {"label": "Cockpit", "value": fact_rows},
                ],
                "checks": checks,
            },
            "table": anomalies[:100],
            "metadata": {
                "engine": "SP2I Data Quality Center",
                "qa_status": "PASS" if not warnings else "WARN",
                "tolerance_capex_pct": 0.005,
                "source": source,
                "metrics": metrics,
                "history": history,
            },
            "warnings": warnings,
        })

    def export_gain_analysis(self, query: AnalyticsQuery) -> BytesIO:
        """Construit un export Excel finance du gain potentiel net."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        analysis = self._build_gain_analysis(query)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Synthese"
        header_fill = PatternFill("solid", fgColor="0B1728")
        header_font = Font(color="FFFFFF", bold=True)

        def write_table(sheet, headers: list[str], rows: list[list[Any]]) -> None:
            sheet.append(headers)
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for row in rows:
                sheet.append(row)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells) + 2
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 14), 42)

        write_table(
            summary,
            ["Indicateur", "Valeur"],
            [
                ["CAPEX local", analysis["kpis"].get("capex_local")],
                ["Cout import final", analysis["kpis"].get("cout_import_final")],
                ["Gain net", analysis["kpis"].get("gain_net")],
                ["Confiance estimation", analysis["kpis"].get("confiance")],
                ["Perimetre", analysis["metadata"].get("scope")],
            ],
        )

        detail = workbook.create_sheet("Detail calcul")
        write_table(
            detail,
            ["Poste", "Montant FCFA", "Sens", "Explication"],
            [
                [row.get("label"), row.get("value"), row.get("direction"), row.get("description")]
                for row in analysis["table"]
            ],
        )

        scenarios = workbook.create_sheet("Scenarios")
        write_table(
            scenarios,
            ["Scenario", "Gain FCFA", "Hypothese"],
            [
                [row.get("label"), row.get("value"), row.get("description")]
                for row in analysis["charts"].get("scenarios", [])
            ],
        )

        risks = workbook.create_sheet("Risques")
        write_table(
            risks,
            ["Risque", "Impact FCFA", "Probabilite", "Action recommandee"],
            [
                [row.get("label"), row.get("impact"), row.get("probabilite"), row.get("action")]
                for row in analysis["charts"].get("risks", [])
            ],
        )

        story = workbook.create_sheet("Storytelling")
        write_table(story, ["Message"], [[line] for line in analysis["metadata"].get("storytelling", [])])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def export_procurement_file(self, query: AnalyticsQuery) -> BytesIO:
        """Dossier achat complet pour audit finance / sourcing Chine."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        gain = self._build_gain_analysis(query)
        suppliers = self._build_suppliers(query)
        scenarios = self._build_procurement_scenarios(query)
        risks = self._build_import_risks(query)
        logistics = self._build_logistics_plan(query)
        table, _ = self.repository.table(query.model_copy(update={"page": 1, "page_size": 5000}))

        workbook = Workbook()
        header_fill = PatternFill("solid", fgColor="0B1728")
        header_font = Font(color="FFFFFF", bold=True)

        def write_table(sheet, headers: list[str], rows: list[list[Any]]) -> None:
            sheet.append(headers)
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for row in rows:
                sheet.append(row)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells) + 2
                sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 14), 46)

        sheet = workbook.active
        sheet.title = "Synthese"
        write_table(
            sheet,
            ["Indicateur", "Valeur"],
            [
                ["Pays sourcing", "Chine"],
                ["Devise active", self._active_currency(query)],
                ["Gain net", gain["kpis"].get("gain_net")],
                ["ROI net", gain["kpis"].get("roi_net")],
                ["Confiance", gain["kpis"].get("confiance")],
                ["Scenario recommande", scenarios["metadata"].get("recommended_scenario")],
            ],
        )

        detail = workbook.create_sheet("Detail gain")
        write_table(detail, ["Poste", "Montant", "Sens", "Explication"], [[row.get("label"), row.get("value"), row.get("direction"), row.get("description")] for row in gain["table"]])

        facts = workbook.create_sheet("Lignes FACT_METRE")
        write_table(
            facts,
            ["Lot", "Designation", "Batiment", "Niveau", "Decision", "CAPEX local", "CAPEX optimise", "Economie"],
            [[row.get("lot"), row.get("designation"), row.get("batiment"), row.get("niveau"), row.get("decision_import"), row.get("capex_local"), row.get("capex_optimise"), row.get("economie")] for row in table],
        )

        suppliers_sheet = workbook.create_sheet("Fournisseurs Chine")
        write_table(
            suppliers_sheet,
            ["Fournisseur", "Categorie", "Port", "FOB USD", "MOQ", "Delai", "Qualite", "Score", "Confiance", "Statut QA", "Certifications", "Source"],
            [[row.get("supplier"), row.get("category"), row.get("port"), row.get("fob_usd"), row.get("moq"), row.get("lead_time_days"), row.get("quality"), row.get("score"), row.get("supplier_confidence_score"), row.get("qa_status"), ", ".join(row.get("certifications", [])), row.get("source_url")] for row in suppliers["table"]],
        )

        containers_sheet = workbook.create_sheet("Containers")
        write_table(containers_sheet, ["Etape", "Delai jours", "Responsable", "Risque"], [[row.get("step"), row.get("days"), row.get("owner"), row.get("risk")] for row in logistics["charts"].get("timeline", [])])

        risks_sheet = workbook.create_sheet("Risques")
        write_table(risks_sheet, ["Risque", "Probabilite", "Impact", "Action"], [[row.get("label"), row.get("probability"), row.get("impact"), row.get("action")] for row in risks["table"]])

        scenarios_sheet = workbook.create_sheet("Scenarios")
        write_table(scenarios_sheet, ["Scenario", "Budget", "ROI", "Qualite", "Risque", "Delai", "Fournisseurs"], [[row.get("label"), row.get("budget"), row.get("roi"), row.get("quality"), row.get("risk"), row.get("lead_time"), ", ".join(row.get("suppliers", []))] for row in scenarios["table"]])

        story = workbook.create_sheet("Storytelling IA")
        write_table(story, ["Message"], [[line] for line in gain["metadata"].get("storytelling", []) + scenarios["metadata"].get("storytelling", [])])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def query_performance(self) -> dict[str, Any]:
        start = perf_counter()
        query = AnalyticsQuery()
        self.repository.kpis(query)
        elapsed_ms = round((perf_counter() - start) * 1000, 2)
        return {
            "status": "SUCCESS",
            "filters": {},
            "pagination": {},
            "kpis": {},
            "charts": {},
            "table": [],
            "metadata": {"last_probe_ms": elapsed_ms, "target_ms": 500},
        }

    def _latest_pipeline_source(self) -> dict[str, Any]:
        chemin_source = RACINE / "03_DONNEES_ENTREE/dqe/dqe_source_brut.json"
        if not chemin_source.exists():
            return {"available": False, "reason": "Aucun fichier source courant trouve."}
        try:
            payload = json.loads(chemin_source.read_text(encoding="utf-8-sig"))
        except Exception as exc:
            return {"available": False, "reason": f"Lecture source impossible: {exc}"}

        audit_excel = payload.get("audit_excel", {}) if isinstance(payload, dict) else {}
        sheet_selection = audit_excel.get("sheet_selection", {})
        ai_preview = audit_excel.get("ai_preview", {})
        lignes = payload.get("lignes", []) if isinstance(payload, dict) else []
        capex_source = 0.0
        for ligne in lignes:
            if not isinstance(ligne, dict):
                continue
            montant = (
                ligne.get("prix_total_ht")
                or ligne.get("montant_total")
                or ligne.get("montant_ht")
                or ligne.get("total_ht")
                or ligne.get("CAPEX_LOCAL")
                or ligne.get("CAPEX_LOCAL_MONTANT")
            )
            try:
                capex_source += float(str(montant or 0).replace(" ", "").replace(",", "."))
            except ValueError:
                quantite = ligne.get("quantite") or 0
                pu = ligne.get("prix_unitaire_ht") or ligne.get("PU_LOCAL") or 0
                try:
                    capex_source += float(quantite or 0) * float(pu or 0)
                except (TypeError, ValueError):
                    pass
        return {
            "available": True,
            "source": payload.get("source") if isinstance(payload, dict) else None,
            "rows_in_source_json": len(lignes),
            "capex_source": round(capex_source, 2),
            "source_fact_metre": sheet_selection.get("source_fact_metre", []),
            "selection_reason": sheet_selection.get("reason"),
            "sheet_evaluations": sheet_selection.get("evaluations", []),
            "ignored_sheets": sheet_selection.get("ignored_sheets", []),
            "blacklist_active": sheet_selection.get("blacklist_active", []),
            "ai_preview": ai_preview,
            "ai_anomalies": audit_excel.get("ai_anomalies", []),
        }

    def _build_dashboard(self, query: AnalyticsQuery, dashboard_type: str) -> dict[str, Any]:
        table, total = self.repository.table(query)
        group_default = {
            "direction": "lot",
            "capex": "lot",
            "procurement": "famille",
            "logistics": "decision_import",
            "chantier": "batiment",
        }.get(dashboard_type, "lot")
        return self._response(
            query,
            kpis=self.repository.kpis(query),
            charts={
                "bar": self.repository.grouped(query, default_group=group_default),
                "heatmap": self.repository.heatmap(query),
                "timeline": self.repository.timeline(query),
                "sankey": self.repository.sankey(query),
            },
            table=table,
            total=total,
            metadata={"dashboard": dashboard_type, "engine": "SP2I Analytics Engine V1"},
        )

    def _build_suppliers(self, query: AnalyticsQuery) -> dict[str, Any]:
        table, _ = self.repository.table(query.model_copy(update={"page": 1, "page_size": 5000}))
        family_gain: dict[str, float] = {}
        family_capex: dict[str, float] = {}
        for row in table:
            family = str(row.get("famille") or row.get("lot") or "").lower()
            for supplier in CHINA_SUPPLIERS:
                if supplier["category"].lower().split()[0] in family or supplier["category"].lower() in str(row.get("lot") or "").lower():
                    family_gain[supplier["category"]] = family_gain.get(supplier["category"], 0) + float(row.get("economie") or 0)
                    family_capex[supplier["category"]] = family_capex.get(supplier["category"], 0) + float(row.get("capex_local") or 0)

        rows: list[dict[str, Any]] = []
        for supplier in CHINA_SUPPLIERS:
            price_score = max(0, 100 - supplier["fob_usd"] / 8)
            delay_score = max(0, 100 - supplier["lead_time_days"] * 0.7)
            moq_score = 82 if "container" not in supplier["moq"].lower() else 68
            score = round(
                price_score * 0.22
                + supplier["quality"] * 0.24
                + delay_score * 0.14
                + moq_score * 0.12
                + supplier["reliability"] * 0.18
                + (100 - supplier["risk"]) * 0.10,
                1,
            )
            supplier_confidence_score = round(
                score * 0.45
                + supplier["quality"] * 0.18
                + supplier["reliability"] * 0.18
                + float(supplier.get("source_confidence") or 0) * 100 * 0.19,
                1,
            )
            gain = family_gain.get(supplier["category"], 0)
            capex = family_capex.get(supplier["category"], 0)
            rows.append({
                **supplier,
                "country": "CN",
                "currency": "USD",
                "score": score,
                "supplier_confidence_score": supplier_confidence_score,
                "estimated_gain": round(gain, 2),
                "capex_scope": round(capex, 2),
                "landed_cost_note": "FOB + maritime + assurance + douane + port + logistique locale + marge securite",
                "qa_status": "A qualifier par RFQ" if supplier_confidence_score < 72 else "Reference exploitable V1",
            })

        rows.sort(key=lambda row: row["score"], reverse=True)
        return self._response(
            query,
            kpis={
                "nb_fournisseurs": len(rows),
                "score_moyen": round(sum(row["score"] for row in rows) / max(len(rows), 1), 1),
                "supplier_confidence_score": round(sum(row["supplier_confidence_score"] for row in rows) / max(len(rows), 1), 1),
                "delai_moyen": round(sum(row["lead_time_days"] for row in rows) / max(len(rows), 1), 1),
                "ports": len(CHINA_PORTS),
            },
            charts={"ports": CHINA_PORTS, "categories": list({row["category"] for row in rows})},
            table=rows,
            total=len(rows),
            metadata={"engine": "SP2I China Supplier Database V1", "country": "CN", "currency": "USD"},
        )

    def _build_procurement_lines(self, query: AnalyticsQuery) -> dict[str, Any]:
        analysis_query = query.model_copy(update={"page": 1, "page_size": 5000})
        table, total = self.repository.table(analysis_query)
        project_kpis = self.repository.kpis(AnalyticsQuery())
        active_currency = self._active_currency(query)
        rows: list[dict[str, Any]] = []

        for row in table:
            lot = normalize_payload_labels({"value": row.get("lot") or ""})["value"]
            famille = normalize_payload_labels({"value": row.get("famille") or ""})["value"]
            designation = normalize_payload_labels({"value": row.get("designation") or ""})["value"]
            capex_local_fcfa = float(row.get("capex_local") or 0)
            capex_import_fcfa = float(row.get("capex_import") or max(capex_local_fcfa - float(row.get("economie") or 0), 0))
            local_supplier = self._match_local_supplier(famille, lot, designation)
            china_supplier = self._match_china_supplier(famille, lot, designation)
            landed = self._line_landed_cost(capex_import_fcfa, china_supplier)
            landed_cost_fcfa = landed["landed_cost"]
            gain_net_fcfa = capex_local_fcfa - landed_cost_fcfa
            roi = gain_net_fcfa / landed_cost_fcfa if landed_cost_fcfa > 0 else 0
            risk_score = min(95, max(10, china_supplier["risk"] + (12 if landed["lead_time_days"] > 60 else 0) - max(roi, 0) * 40))
            confidence = min(96, max(45, china_supplier.get("source_confidence", 0.55) * 100 * 0.46 + china_supplier["quality"] * 0.26 + local_supplier["quality"] * 0.14 + max(0, min(roi * 180, 14))))

            if roi >= 0.12 and risk_score < 58:
                decision = "IMPORT"
            elif roi >= 0.06 and risk_score < 72:
                decision = "HYBRIDE"
            elif roi > 0 and risk_score >= 72:
                decision = "A ETUDIER"
            else:
                decision = "LOCAL"
            reasons = self._line_decision_reasons(decision, roi, risk_score, china_supplier, local_supplier, landed)

            rows.append({
                "id_ligne": row.get("id_ligne"),
                "designation": designation,
                "quantite": float(row.get("quantite") or 0),
                "unite": row.get("unite") or "u",
                "lot": lot,
                "famille": famille,
                "fournisseur_local": local_supplier["supplier"],
                "pays_local": f"{local_supplier['city']}, {local_supplier['country']}",
                "delai_local": local_supplier["lead_time_days"],
                "qualite_locale": local_supplier["quality"],
                "historique_local": local_supplier["history"],
                "prix_local": round(self._from_fcfa(capex_local_fcfa, active_currency), 2),
                "fournisseur_chine": china_supplier["supplier"],
                "province_chine": china_supplier["port"],
                "port_chine": china_supplier["port"],
                "moq_chine": china_supplier["moq"],
                "certifications_chine": ", ".join(china_supplier["certifications"]),
                "score_fournisseur_chine": china_supplier["quality"],
                "fob_chine": round(self._from_fcfa(landed["fob"], active_currency), 2),
                "landed_cost_chine": round(self._from_fcfa(landed_cost_fcfa, active_currency), 2),
                "gain_net": round(self._from_fcfa(gain_net_fcfa, active_currency), 2),
                "roi_import": round(roi, 5),
                "risque": round(risk_score, 1),
                "delai": landed["lead_time_days"],
                "decision_ia": decision,
                "score_confiance_ia": round(confidence, 1),
                "decision_reasons": reasons,
                "storytelling": self._line_storytelling(decision, china_supplier, gain_net_fcfa, roi, risk_score),
                "landed_cost_detail": landed,
                "containers": max(1, round(landed_cost_fcfa / 42_000_000, 2)),
                "currency": active_currency,
            })

        import_rows = [row for row in rows if row["decision_ia"] == "IMPORT"]
        hybrid_rows = [row for row in rows if row["decision_ia"] == "HYBRIDE"]
        capex_local_total = sum(float(row.get("prix_local") or 0) for row in rows)
        landed_total = sum(float(row.get("landed_cost_chine") or 0) for row in rows)
        project_capex_total = self._from_fcfa(float(project_kpis.get("capex_brut") or 0), active_currency)
        logistics_cost_total = sum(self._from_fcfa(float(row.get("landed_cost_detail", {}).get("maritime") or 0), active_currency) for row in rows)
        customs_cost_total = sum(self._from_fcfa(float(row.get("landed_cost_detail", {}).get("douane") or 0), active_currency) for row in rows)
        containers_needed = round(sum(float(row.get("containers") or 0) for row in rows), 1)
        main_family = query.filters.famille or query.filters.lot or (rows[0]["famille"] if rows else "Famille selectionnee")
        positive_roi_rows = [row for row in rows if float(row.get("roi_import") or 0) > 0]
        top_gain = sorted(rows, key=lambda item: item["gain_net"], reverse=True)[:3]
        decision_breakdown = {
            decision: len([row for row in rows if row["decision_ia"] == decision])
            for decision in ["IMPORT", "HYBRIDE", "A ETUDIER", "LOCAL"]
        }
        comparison = {
            "local": {
                "cost": round(capex_local_total, 2),
                "lead_time": round(sum(float(row.get("delai_local") or 0) for row in rows) / max(len(rows), 1), 1),
                "quality": round(sum(float(row.get("qualite_locale") or 0) for row in rows) / max(len(rows), 1), 1),
                "risk": 32,
                "availability": "Forte proximite chantier",
            },
            "china": {
                "cost": round(landed_total, 2),
                "lead_time": round(sum(float(row.get("delai") or 0) for row in rows) / max(len(rows), 1), 1),
                "quality": round(sum(float(row.get("score_fournisseur_chine") or 0) for row in rows) / max(len(rows), 1), 1),
                "risk": round(sum(float(row.get("risque") or 0) for row in rows) / max(len(rows), 1), 1),
                "availability": "MOQ et consolidation container",
            },
        }
        timeline = [
            {"step": "Commande", "days": 3, "risk": "Faible"},
            {"step": "Fabrication", "days": 32, "risk": "Moyen"},
            {"step": "Container", "days": 7, "risk": "Moyen"},
            {"step": "Maritime", "days": 28, "risk": "Moyen"},
            {"step": "Douane", "days": 8, "risk": "Eleve"},
            {"step": "Chantier", "days": 3, "risk": "Faible"},
        ]
        storytelling = self._family_storytelling(
            main_family,
            capex_local_total,
            project_capex_total,
            positive_roi_rows,
            rows,
            top_gain,
            decision_breakdown,
        )
        return self._response(
            query,
            kpis={
                "nb_lignes": len(rows),
                "nb_import": len(import_rows),
                "nb_hybride": len(hybrid_rows),
                "gain_net_total": round(sum(row["gain_net"] for row in rows), 2),
                "roi_moyen": round(sum(row["roi_import"] for row in rows) / max(len(rows), 1), 5),
                "risque_moyen": round(sum(row["risque"] for row in rows) / max(len(rows), 1), 1),
                "capex_local": round(capex_local_total, 2),
                "capex_chine_rendu_chantier": round(landed_total, 2),
                "cout_logistique": round(logistics_cost_total, 2),
                "cout_douane": round(customs_cost_total, 2),
                "containers": containers_needed,
                "part_capex_projet": round(capex_local_total / project_capex_total, 6) if project_capex_total else 0,
                "nb_fournisseurs": len({row["fournisseur_local"] for row in rows} | {row["fournisseur_chine"] for row in rows}),
                "devise": active_currency,
            },
            charts={
                "decisions": [
                    {"label": decision, "value": len([row for row in rows if row["decision_ia"] == decision])}
                    for decision in ["IMPORT", "HYBRIDE", "A ETUDIER", "LOCAL"]
                ],
                "top_gain": sorted(rows, key=lambda item: item["gain_net"], reverse=True)[:12],
                "comparison": comparison,
                "timeline": timeline,
                "containers": {
                    "count": containers_needed,
                    "cbm": round(containers_needed * 58, 1),
                    "mutualisation": "Regrouper les lignes IMPORT et HYBRIDE sur containers partages.",
                    "logistics_cost": round(logistics_cost_total, 2),
                },
                "waterfall": self._family_waterfall(rows),
            },
            table=rows,
            total=total,
            metadata={
                "engine": "SP2I Family Procurement Cockpit V1",
                "currency": active_currency,
                "country": "CN",
                "family_scope": main_family,
                "project_capex_total": round(project_capex_total, 2),
                "storytelling": storytelling,
                "decision_breakdown": decision_breakdown,
            },
        )

    def _build_procurement_scenarios(self, query: AnalyticsQuery) -> dict[str, Any]:
        gain = self._build_gain_analysis(query)
        capex_local = float(gain["kpis"].get("capex_local") or 0)
        active_currency = self._active_currency(query)
        rows = []
        for scenario in CHINA_SCENARIOS:
            estimated_budget = capex_local * (1 - scenario["discount"] + scenario["security"])
            gain_net = max(capex_local - estimated_budget, 0)
            rows.append({
                "code": scenario["code"],
                "label": scenario["label"],
                "budget": round(self._from_fcfa(estimated_budget, active_currency), 2),
                "gain_net": round(self._from_fcfa(gain_net, active_currency), 2),
                "roi": round(gain_net / estimated_budget, 5) if estimated_budget else 0,
                "quality": scenario["quality"],
                "risk": scenario["risk"],
                "lead_time": scenario["lead_time"],
                "suppliers": [supplier["supplier"] for supplier in CHINA_SUPPLIERS if supplier["quality"] >= scenario["quality"] - 6][:3],
                "currency": active_currency,
                "landed_cost": round(self._from_fcfa(estimated_budget, active_currency), 2),
            })

        recommended = sorted(rows, key=lambda row: (row["roi"], -row["risk"], row["quality"]), reverse=True)[0]
        storytelling = [
            f"{recommended['label']} offre le meilleur equilibre gain / risque pour le perimetre courant.",
            "Eco China maximise le prix, Premium China protege la qualite et Standard China sert de scenario de reference.",
            "Les prix restent en devise source USD et sont convertis pour le cockpit sans ecraser la devise d'origine.",
        ]
        return self._response(
            query,
            kpis={"recommended_roi": recommended["roi"], "recommended_budget": recommended["budget"], "nb_scenarios": len(rows)},
            charts={"scenarios": rows},
            table=rows,
            total=len(rows),
            metadata={"engine": "SP2I Procurement Scenario Engine V1", "recommended_scenario": recommended["label"], "storytelling": storytelling},
        )

    def _build_import_risks(self, query: AnalyticsQuery) -> dict[str, Any]:
        kpis = self.repository.kpis(query)
        capex_import = float(kpis.get("capex_optimise") or 0)
        rows = [
            {
                **risk,
                "impact": round(capex_import * risk["impact_rate"], 2),
                "criticite": round((risk["probability"] * 100) * 0.55 + min(risk["impact_rate"] * 1800, 45), 1),
            }
            for risk in IMPORT_RISKS
        ]
        rows.sort(key=lambda row: row["criticite"], reverse=True)
        return self._response(
            query,
            kpis={
                "risque_global": round(sum(row["criticite"] for row in rows) / max(len(rows), 1), 1),
                "impact_total": round(sum(row["impact"] for row in rows), 2),
                "nb_risques": len(rows),
            },
            charts={"risk_map": rows},
            table=rows,
            total=len(rows),
            metadata={"engine": "SP2I China Import Risk Engine V1"},
        )

    def _build_logistics_plan(self, query: AnalyticsQuery) -> dict[str, Any]:
        kpis = self.repository.kpis(query)
        capex = float(kpis.get("capex_optimise") or 0)
        container_capacity = 42_000_000
        containers = max(1, round(capex / container_capacity))
        timeline = [
            {"step": "Commande fournisseur", "days": 3, "owner": "Achats", "risk": "Faible"},
            {"step": "Fabrication Chine", "days": 32, "owner": "Fournisseur", "risk": "Moyen"},
            {"step": "Consolidation container", "days": 7, "owner": "Transitaire", "risk": "Moyen"},
            {"step": "Maritime", "days": 28, "owner": "Armateur", "risk": "Moyen"},
            {"step": "Douane et port", "days": 8, "owner": "Declarant", "risk": "Eleve"},
            {"step": "Livraison chantier", "days": 3, "owner": "Logistique locale", "risk": "Faible"},
        ]
        return self._response(
            query,
            kpis={"containers_estimes": containers, "cbm_estime": containers * 58, "delai_total": sum(row["days"] for row in timeline)},
            charts={"timeline": timeline, "ports": CHINA_PORTS},
            table=timeline,
            total=len(timeline),
            metadata={"engine": "SP2I China Logistics Engine V1", "container_capacity_fcfa": container_capacity},
        )

    def _build_gain_analysis(self, query: AnalyticsQuery) -> dict[str, Any]:
        analysis_query = query.model_copy(update={"page": 1, "page_size": 5000})
        table, total = self.repository.table(analysis_query)
        kpis = self.repository.kpis(query)

        capex_local = sum(float(row.get("capex_local") or row.get("capex_brut") or 0) for row in table)
        capex_import_recorded = sum(float(row.get("capex_import") or 0) for row in table)
        economie_recorded = sum(float(row.get("economie") or row.get("economie_nette") or 0) for row in table)
        if not capex_local:
            capex_local = float(kpis.get("capex_brut") or 0)
        if not capex_import_recorded:
            capex_import_recorded = max(capex_local - float(kpis.get("economie_nette") or 0), 0)
        if not economie_recorded:
            economie_recorded = float(kpis.get("economie_nette") or 0)

        import_rows = [row for row in table if str(row.get("decision_import") or "").upper() == "IMPORT"]
        import_share = len(import_rows) / len(table) if table else float(kpis.get("taux_importable") or 0)
        landed_rates = {
            "transport": 0.075,
            "douane": 0.052,
            "assurance": 0.014,
            "logistique": 0.025,
            "marge_securite": 0.018,
            "risques_estimes": 0.022,
        }
        landed_rate_total = sum(landed_rates.values())

        # Les donnees FACT_METRE stockent un cout import exploitable cockpit. Pour
        # rendre le calcul audit-proof, on le recompose en FOB + couts rendus chantier.
        cout_import_final = max(capex_import_recorded, capex_local - economie_recorded)
        capex_import_fob = cout_import_final / (1 + landed_rate_total) if cout_import_final else 0
        transport = capex_import_fob * landed_rates["transport"]
        douane = capex_import_fob * landed_rates["douane"]
        assurance = capex_import_fob * landed_rates["assurance"]
        logistique = capex_import_fob * landed_rates["logistique"]
        marge_securite = capex_import_fob * landed_rates["marge_securite"]
        risques_estimes = capex_import_fob * landed_rates["risques_estimes"]
        cout_import_final = capex_import_fob + transport + douane + assurance + logistique + marge_securite + risques_estimes
        gain_net = max(capex_local - cout_import_final, 0)
        gain_brut_achat = max(capex_local - capex_import_fob, 0)
        roi_net = gain_net / cout_import_final if cout_import_final else 0

        quality = self.repository.quality_metrics()
        taxonomy_gap = float(quality.get("lignes_famille_a_classer") or 0) / max(float(quality.get("nb_lignes") or 1), 1)
        confidence = 0.88 - min(taxonomy_gap * 0.25, 0.12) - min(import_share * 0.08, 0.08)
        confidence = round(max(0.62, min(0.94, confidence)), 2)

        risks = [
            {
                "label": "Hausse transport",
                "impact": round(-(transport * 0.35), 2),
                "probabilite": "Moyenne",
                "action": "Verrouiller le fret et mutualiser les containers.",
            },
            {
                "label": "Retard douane",
                "impact": round(-(douane * 0.25), 2),
                "probabilite": "Moyenne",
                "action": "Anticiper documents, HS codes et inspection.",
            },
            {
                "label": "Variation devise",
                "impact": round(-(capex_import_fob * 0.03), 2),
                "probabilite": "Elevee",
                "action": "Fixer les devis ou ajouter une clause de change.",
            },
            {
                "label": "Qualite fournisseur",
                "impact": round(-(capex_import_fob * 0.02), 2),
                "probabilite": "Faible",
                "action": "Prevoir controle qualite avant embarquement.",
            },
        ]

        scenarios = [
            {
                "label": "Optimiste",
                "value": round(gain_net * 1.28, 2),
                "description": "Fret stable, douane fluide, prix fournisseur confirme.",
            },
            {
                "label": "Realiste",
                "value": round(gain_net, 2),
                "description": "Hypothese SP2I avec marge securite et risques integres.",
            },
            {
                "label": "Pessimiste",
                "value": round(max(gain_net + sum(risk["impact"] for risk in risks), 0), 2),
                "description": "Transport et devise defavorables, retards limites.",
            },
        ]

        decomposition = [
            {"label": "CAPEX local", "value": round(capex_local, 2), "direction": "base", "description": "Budget si les achats restent en local."},
            {"label": "CAPEX import FOB", "value": round(capex_import_fob, 2), "direction": "cost", "description": "Prix fournisseur avant couts rendus chantier."},
            {"label": "Transport", "value": round(transport, 2), "direction": "cost", "description": "Fret maritime et consolidation logistique."},
            {"label": "Douane", "value": round(douane, 2), "direction": "cost", "description": "Droits, taxes et formalites import."},
            {"label": "Assurance", "value": round(assurance, 2), "direction": "cost", "description": "Couverture transport et marchandises."},
            {"label": "Logistique chantier", "value": round(logistique, 2), "direction": "cost", "description": "Port, manutention et livraison chantier."},
            {"label": "Marge securite", "value": round(marge_securite, 2), "direction": "risk", "description": "Buffer finance pour ecarts de cotation."},
            {"label": "Risques estimes", "value": round(risques_estimes, 2), "direction": "risk", "description": "Impact attendu des risques import."},
            {"label": "Gain net final", "value": round(gain_net, 2), "direction": "gain", "description": "Economie apres tous les couts et risques."},
        ]

        waterfall = [
            {"label": "CAPEX local", "value": round(capex_local, 2), "type": "total"},
            {"label": "Gains achat", "value": round(-gain_brut_achat, 2), "type": "gain"},
            {"label": "Transport", "value": round(transport, 2), "type": "cost"},
            {"label": "Douane", "value": round(douane, 2), "type": "cost"},
            {"label": "Assurance", "value": round(assurance, 2), "type": "cost"},
            {"label": "Logistique", "value": round(logistique, 2), "type": "cost"},
            {"label": "Risques", "value": round(marge_securite + risques_estimes, 2), "type": "risk"},
            {"label": "Gain net", "value": round(gain_net, 2), "type": "final"},
        ]

        scope = (
            query.filters.lot
            or query.filters.famille
            or query.filters.batiment
            or query.filters.decision_import
            or "Portefeuille achats global"
        )
        active_currency = self._active_currency(query)
        convert_money = lambda value: round(self._from_fcfa(float(value or 0), active_currency), 2)
        display_decomposition = [{**row, "value": convert_money(row.get("value"))} for row in decomposition]
        display_waterfall = [{**row, "value": convert_money(row.get("value"))} for row in waterfall]
        display_scenarios = [{**row, "value": convert_money(row.get("value"))} for row in scenarios]
        display_risks = [{**row, "impact": convert_money(row.get("impact"))} for row in risks]
        storytelling = [
            f"Le gain net estime est de {convert_money(gain_net):,.0f} {active_currency} apres transport, douane, assurance, logistique et risques.",
            f"La confiance d'estimation est de {round(confidence * 100)} %, basee sur la qualite du dataset, la volatilite import et le taux importable.",
            "La recommandation SP2I est de valider l'import lorsque le gain net reste positif dans le scenario pessimiste.",
        ]

        return self._response(
            query,
            kpis={
                "capex_local": convert_money(capex_local),
                "capex_import_fob": convert_money(capex_import_fob),
                "cout_import_final": convert_money(cout_import_final),
                "gain_brut_achat": convert_money(gain_brut_achat),
                "gain_net": convert_money(gain_net),
                "roi_net": round(roi_net, 6),
                "confiance": confidence,
                "scenario_optimiste": display_scenarios[0]["value"],
                "scenario_realiste": display_scenarios[1]["value"],
                "scenario_pessimiste": display_scenarios[2]["value"],
                "nb_lignes": total,
                "devise": active_currency,
            },
            charts={
                "waterfall": display_waterfall,
                "scenarios": display_scenarios,
                "risks": display_risks,
                "sensitivity": {
                    "min": display_scenarios[2]["value"],
                    "max": display_scenarios[0]["value"],
                    "realiste": display_scenarios[1]["value"],
                },
            },
            table=display_decomposition,
            total=len(decomposition),
            metadata={
                "engine": "SP2I Gain Potential Engine V1",
                "scope": scope,
                "currency": active_currency,
                "formula": "Gain net = CAPEX local - (CAPEX import FOB + transport + douane + assurance + logistique + marge securite + risques estimes)",
                "assumptions": landed_rates,
                "storytelling": storytelling,
                "source": "FACT_METRE",
            },
        )

    def _response(
        self,
        query: AnalyticsQuery,
        kpis: dict[str, Any] | None = None,
        charts: dict[str, Any] | None = None,
        table: list[dict[str, Any]] | None = None,
        total: int | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        return normalize_payload_labels({
            "status": "SUCCESS",
            "filters": query.filters.model_dump(exclude_none=True),
            "pagination": {
                "page": query.page,
                "page_size": query.page_size,
                "total": total if total is not None else len(table or []),
            },
            "kpis": kpis or {},
            "charts": charts or {},
            "table": table or [],
            "metadata": metadata or {"engine": "SP2I Analytics Engine V1"},
        })

    def _cached(self, prefix: str, query: AnalyticsQuery, builder) -> dict[str, Any]:
        key = f"{prefix}:{hashlib.sha1(query.model_dump_json().encode()).hexdigest()}"
        cached = analytics_cache.get(key)
        if cached:
            cached["metadata"] = {**cached.get("metadata", {}), "cache_hit": True}
            return cached
        value = builder()
        value["metadata"] = {**value.get("metadata", {}), "cache_hit": False}
        analytics_cache.set(key, json.loads(json.dumps(value, default=str)))
        return value

    @staticmethod
    def _active_currency(query: AnalyticsQuery) -> str:
        currency = str(query.filters.devise or "FCFA").upper()
        return currency if currency in SUPPORTED_CURRENCIES else "FCFA"

    @staticmethod
    def _from_fcfa(value: float, currency: str) -> float:
        rate = SUPPORTED_CURRENCIES.get(currency, SUPPORTED_CURRENCIES["FCFA"])["to_fcfa"]
        return value / rate if rate else value

    @staticmethod
    def _text_match_score(text: str, candidates: list[str]) -> int:
        lowered = text.lower()
        return sum(1 for candidate in candidates if candidate and candidate.lower() in lowered)

    def _match_china_supplier(self, famille: str, lot: str, designation: str) -> dict[str, Any]:
        text = f"{famille} {lot} {designation}".lower()
        category_keywords = {
            "Alucobond": ["alucobond", "facade", "habillage"],
            "Electricite": ["electric", "cable", "courant", "tableau", "luminaire"],
            "CVC": ["cvc", "clim", "cta", "ventilation", "extraction"],
            "Sanitaires": ["plomberie", "sanitaire", "evacuation", "eau"],
            "Menuiserie aluminium": ["menuiserie", "aluminium", "vitrerie", "baie"],
            "Ascenseurs": ["ascenseur", "lift"],
            "Mobilier medical": ["medical", "mobilier", "lit", "paillasse"],
        }
        best = CHINA_SUPPLIERS[0]
        best_score = -1
        for supplier in CHINA_SUPPLIERS:
            score = self._text_match_score(text, category_keywords.get(supplier["category"], []))
            score += 2 if supplier["category"].lower() in text else 0
            if score > best_score:
                best = supplier
                best_score = score
        return best

    def _match_local_supplier(self, famille: str, lot: str, designation: str) -> dict[str, Any]:
        text = f"{famille} {lot} {designation}".lower()
        for key, supplier in LOCAL_SUPPLIERS.items():
            if key != "default" and key in text:
                return supplier
        if "electric" in text:
            return LOCAL_SUPPLIERS["electricite"]
        if "clim" in text or "cvc" in text:
            return LOCAL_SUPPLIERS["cvc"]
        if "sanitaire" in text or "plomberie" in text:
            return LOCAL_SUPPLIERS["plomberie"]
        if "aluminium" in text or "menuiserie" in text:
            return LOCAL_SUPPLIERS["menuiserie"]
        return LOCAL_SUPPLIERS["default"]

    @staticmethod
    def _line_landed_cost(import_cost_fcfa: float, supplier: dict[str, Any]) -> dict[str, Any]:
        base = max(import_cost_fcfa, supplier["fob_usd"] * SUPPORTED_CURRENCIES["USD"]["to_fcfa"])
        fob = base / 1.205
        maritime = fob * 0.075
        assurance = fob * 0.014
        douane = fob * 0.052
        logistique_locale = fob * 0.025
        marge_securite = fob * 0.018
        landed_cost = fob + maritime + assurance + douane + logistique_locale + marge_securite
        return {
            "fob": round(fob, 2),
            "maritime": round(maritime, 2),
            "assurance": round(assurance, 2),
            "douane": round(douane, 2),
            "logistique_locale": round(logistique_locale, 2),
            "marge_securite": round(marge_securite, 2),
            "landed_cost": round(landed_cost, 2),
            "lead_time_days": int(supplier["lead_time_days"] + 28),
        }

    @staticmethod
    def _line_storytelling(decision: str, supplier: dict[str, Any], gain_fcfa: float, roi: float, risk_score: float) -> str:
        if decision == "IMPORT":
            return (
                f"Le fournisseur Chine {supplier['supplier']} permet un gain net estime a {round(roi * 100)} %. "
                f"Le principal risque concerne le delai maritime et un score risque de {round(risk_score)}/100."
            )
        if decision == "HYBRIDE":
            return (
                f"Une strategie hybride est recommandee : importer la partie standardisable via {supplier['supplier']} "
                "et conserver en local les besoins urgents ou sensibles au planning."
            )
        if decision == "A ETUDIER":
            return (
                f"Le gain existe mais reste sensible au risque fournisseur/logistique. Lancer une RFQ et verifier certifications {', '.join(supplier['certifications'][:2])}."
            )
        return "Le fournisseur local reste preferable pour proteger le delai chantier ou lorsque le gain net import n'est pas suffisant."

    @staticmethod
    def _line_decision_reasons(
        decision: str,
        roi: float,
        risk_score: float,
        china_supplier: dict[str, Any],
        local_supplier: dict[str, Any],
        landed: dict[str, Any],
    ) -> list[dict[str, str]]:
        reasons: list[dict[str, str]] = []
        if roi > 0.15:
            reasons.append({"type": "positive", "label": "ROI superieur a 15%"})
        elif roi > 0.06:
            reasons.append({"type": "positive", "label": "ROI positif mais a securiser"})
        else:
            reasons.append({"type": "warning", "label": "Gain import insuffisant"})

        if china_supplier.get("source_confidence", 0) >= 0.75:
            reasons.append({"type": "positive", "label": "Fournisseur Chine reference V1"})
        else:
            reasons.append({"type": "warning", "label": "Fournisseur Chine a qualifier par RFQ"})

        if risk_score >= 70:
            reasons.append({"type": "warning", "label": "Risque import eleve"})
        elif risk_score <= 55:
            reasons.append({"type": "positive", "label": "Risque import maitrisable"})

        if landed["lead_time_days"] > 60:
            reasons.append({"type": "warning", "label": f"Delai maritime estime {landed['lead_time_days']} jours"})
        if decision == "LOCAL":
            reasons.append({"type": "positive", "label": f"Delai local court : {local_supplier['lead_time_days']} jours"})
        if decision == "HYBRIDE":
            reasons.append({"type": "neutral", "label": "Standardiser les volumes import et garder le critique en local"})
        return reasons

    @staticmethod
    def _family_waterfall(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        capex_local = sum(float(row.get("prix_local") or 0) for row in rows)
        fob = sum(float(row.get("fob_chine") or 0) for row in rows)
        maritime = sum(float(row.get("landed_cost_detail", {}).get("maritime") or 0) for row in rows)
        douane = sum(float(row.get("landed_cost_detail", {}).get("douane") or 0) for row in rows)
        assurance = sum(float(row.get("landed_cost_detail", {}).get("assurance") or 0) for row in rows)
        logistique = sum(float(row.get("landed_cost_detail", {}).get("logistique_locale") or 0) for row in rows)
        marge = sum(float(row.get("landed_cost_detail", {}).get("marge_securite") or 0) for row in rows)
        landed = sum(float(row.get("landed_cost_chine") or 0) for row in rows)
        return [
            {"label": "CAPEX local", "value": round(capex_local, 2), "type": "total"},
            {"label": "FOB Chine", "value": round(-(capex_local - fob), 2), "type": "gain"},
            {"label": "Maritime", "value": round(maritime, 2), "type": "cost"},
            {"label": "Douane", "value": round(douane, 2), "type": "cost"},
            {"label": "Assurance", "value": round(assurance, 2), "type": "cost"},
            {"label": "Logistique", "value": round(logistique, 2), "type": "cost"},
            {"label": "Risques", "value": round(marge, 2), "type": "risk"},
            {"label": "Gain final", "value": round(max(capex_local - landed, 0), 2), "type": "final"},
        ]

    @staticmethod
    def _family_storytelling(
        family: str,
        family_capex: float,
        project_capex: float,
        positive_roi_rows: list[dict[str, Any]],
        rows: list[dict[str, Any]],
        top_gain: list[dict[str, Any]],
        decision_breakdown: dict[str, int],
    ) -> list[str]:
        share = family_capex / project_capex if project_capex else 0
        positive_rate = len(positive_roi_rows) / len(rows) if rows else 0
        top_labels = ", ".join(row.get("designation", "ligne")[:42] for row in top_gain if row.get("gain_net", 0) > 0) or "les lignes importables les plus standardisees"
        return [
            f"La famille {family} represente {round(share * 100, 1)}% du CAPEX projet, calcule sur le budget projet complet et non sur la vue filtree.",
            f"{round(positive_rate * 100)}% des lignes presentent un ROI import positif apres cout rendu chantier.",
            f"Les principaux gains proviennent de : {top_labels}.",
            f"Repartition IA : {decision_breakdown.get('IMPORT', 0)} IMPORT, {decision_breakdown.get('HYBRIDE', 0)} HYBRIDE, {decision_breakdown.get('LOCAL', 0)} LOCAL, {decision_breakdown.get('A ETUDIER', 0)} a etudier.",
            "Les composants critiques chantier restent recommandes en LOCAL ou HYBRIDE lorsque le delai maritime fragilise le planning.",
        ]
